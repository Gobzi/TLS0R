import os
import subprocess
import pandas as pd
import glob
import argparse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

parser = argparse.ArgumentParser(description='TLS scan and spreadsheet generation script.')
parser.add_argument('--scan', action='store_true', help='Run a scan using testssl.sh on the targets specified in targets.txt.')
parser.add_argument('--generate', action='store_true', help='Generate a spreadsheet from existing CSV files.')
args = parser.parse_args()

testssl_dir = 'testssl.sh'
testssl_script = os.path.join(testssl_dir, 'testssl.sh')

if not os.path.isfile('testssl.sh') and not os.path.isfile(testssl_script):
    print("testssl.sh not found, cloning the repo...")
    subprocess.run(["git", "clone", "https://github.com/drwetter/testssl.sh"])
    print("Repo cloned.")
else:
    print("testssl.sh found.")

csvs_directory = 'csvs'
if not os.path.exists(csvs_directory):
    os.makedirs(csvs_directory)
    print("CSV directory created.")
else:
    print("CSV directory found.")

if args.scan:
    targets_file = 'targets.txt'
    if not os.path.isfile(targets_file):
        print("targets.txt is required.")
        exit(1)
    else:
        print("Targets loaded.")
        with open(targets_file, 'r') as f:
            targets = f.readlines()

        if not targets:
            print("No targets found in targets.txt")
            exit(1)

        for target in targets:
            target = target.strip()
            if target:
                output_file = os.path.join(csvs_directory, f"{target.replace(':', '_').replace('/', '_')}.csv")
                subprocess.run([testssl_script, "--csvfile", output_file, target])
        print("Scanning complete.")

if args.generate:
    os.chdir(csvs_directory)

    allowed_ids_mapping = {
        'SSLv2': 'SSL v2.0',
        'SSLv3': 'SSL v3.0',
        'TLS1': 'TLS v1.0',
        'TLS1_1': 'TLS v1.1',
        'heartbleed': 'Heartbleed',
        'CCS': 'CCS',
        'ticketbleed': 'Ticketbleed',
        'ROBOT': 'ROBOT',
        'secure_renego': 'Secure Renegotiation',
        'secure_client_renego': 'Secure Client Renegotiation',
        'CRIME_TLS': 'CRIME (TLS)',
        'BREACH': 'BREACH',
        'POODLE_SSL': 'POODLE (SSL)',
        'fallback_SCSV': 'Fallback SCSV',
        'SWEET32': 'SWEET32',
        'FREAK': 'FREAK',
        'DROWN': 'DROWN',
        'LOGJAM-common_primes': 'LOGJAM (Common Primes)',
        'LOGJAM': 'LOGJAM',
        'BEAST_CBC_SSL3': 'BEAST (CBC SSL v3.0)',
        'BEAST_CBC_TLS1': 'BEAST (CBC TLS v1.0)',
        'BEAST': 'BEAST',
        'LUCKY13': 'LUCKY13',
        'winshock': 'Winshock',
        'RC4': 'RC4'
    }

    weak_cipher_suites_id = [
        'cipherlist_NULL', 'cipherlist_aNULL', 'cipherlist_EXPORT', 'cipherlist_LOW',
        'cipherlist_3DES_IDEA', 'cipherlist_OBSOLETED', 'cipherlist_STRONG_NOFS',
        'cipherlist_STRONG_FS'
    ]

    certificate_issues_id = [
        'cert_numbers', 'cert_signatureAlgorithm', 'cert_keySize', 'cert_keyUsage',
        'cert_extKeyUsage', 'cert_serialNumber', 'cert_serialNumberLen', 'cert_fingerprintSHA1',
        'cert_fingerprintSHA256', 'cert', 'cert_commonName', 'cert_commonName_wo_SNI',
        'cert_subjectAltName', 'cert_trust', 'cert_chain_of_trust', 'cert_certificatePolicies_EV',
        'cert_expirationStatus', 'cert_notBefore', 'cert_notAfter', 'cert_extlifeSpan',
        'cert_eTLS', 'cert_crlDistributionPoints', 'cert_ocspURL', 'OCSP_stapling',
        'cert_ocspRevoked', 'cert_mustStapleExtension', 'DNS_CAArecord', 'certificate_transparency',
        'certs_countServer', 'certs_list_ordering_problem', 'cert_caIssuers',
        'intermediate_cert', 'intermediate_cert_fingerprintSHA256', 'intermediate_cert_notBefore',
        'intermediate_cert_notAfter', 'intermediate_cert_expiration', 'intermediate_cert_chain'
    ]

    HSTS_id = [
        'HSTS', '', 'HSTS_preload', 'HSTS_time',
        'HSTS_subdomains'
    ]

    allowed_ids_mapping['weak_cipher_suites'] = 'Weak Cipher Suites'
    allowed_ids_mapping['certificate_issues'] = 'Certificate Issues'
    allowed_ids_mapping['HSTS'] = 'HSTS Issues'

    header_font = Font(name='Arial', bold=True, size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')

    def style_cell(cell, condition):
        cell.font = Font(name='Arial')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if condition == 'Vulnerable':
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(name='Arial', bold=True)
        elif condition == 'Secure':
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = 'TLS Report'
    headers = ['IP/Port'] + [allowed_ids_mapping[id] for id in allowed_ids_mapping.keys()]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment

    findings_dict = {}
    weak_cipher_ids = set(weak_cipher_suites_id)
    certificate_issues_ids = set(weak_cipher_suites_id)
    HSTS_ids = set(HSTS_id)
    csv_files = glob.glob(os.path.join('.', '*.csv'))
    for csv_file_path in csv_files:
        df = pd.read_csv(csv_file_path)
        filtered_df = df[~df['severity'].isin(['OK', 'INFO'])]
        for _, row in filtered_df.iterrows():
            if row['id'] in allowed_ids_mapping or row['id'] in weak_cipher_ids:
                ip = row['fqdn/ip'].split('/')[0]
                ip_port = f"{ip}:{row['port']}"
                if ip_port not in findings_dict:
                    findings_dict[ip_port] = {allowed_ids_mapping.get(id, 'Secure'): 'Secure' for id in allowed_ids_mapping}
                if row['id'] in weak_cipher_ids:
                    findings_dict[ip_port]['Weak Cipher Suites'] = 'Vulnerable'
                else:
                    findings_dict[ip_port][allowed_ids_mapping[row['id']]] = 'Vulnerable'

        for _, row in filtered_df.iterrows():
            if row['id'] in allowed_ids_mapping or row['id'] in certificate_issues_ids:
                ip = row['fqdn/ip'].split('/')[0]
                ip_port = f"{ip}:{row['port']}"
                if ip_port not in findings_dict:
                    findings_dict[ip_port] = {allowed_ids_mapping.get(id, 'Secure'): 'Secure' for id in allowed_ids_mapping}
                if row['id'] in certificate_issues_ids:
                    findings_dict[ip_port]['Certificate Issues'] = 'Vulnerable'
                else:
                    findings_dict[ip_port][allowed_ids_mapping[row['id']]] = 'Vulnerable'

        for _, row in filtered_df.iterrows():
            if row['id'] in allowed_ids_mapping or row['id'] in HSTS_ids:
                ip = row['fqdn/ip'].split('/')[0]
                ip_port = f"{ip}:{row['port']}"
                if ip_port not in findings_dict:
                    findings_dict[ip_port] = {allowed_ids_mapping.get(id, 'Secure'): 'Secure' for id in allowed_ids_mapping}
                if row['id'] in certificate_issues_ids:
                    findings_dict[ip_port]['HSTS Issues'] = 'Vulnerable'
                else:
                    findings_dict[ip_port][allowed_ids_mapping[row['id']]] = 'Vulnerable'

    for ip_port, findings in findings_dict.items():
        row_data = [ip_port] + [findings.get(formatted_id, 'Secure') for formatted_id in headers[1:]]
        ws.append(row_data)
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.column == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Arial')
            else:
                style_cell(cell, cell.value)
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    output_file_path = "../TLS_report.xlsx"
    wb.save(output_file_path)
    print(f"TLS Report Spreadsheet generated: TLS_report.xlsx")
